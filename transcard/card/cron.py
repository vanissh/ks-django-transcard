import datetime
import os
import xlwt
import openpyxl
from zipfile import ZipFile

from .models import Card, Loadout


def makedirs(data, media, catalogs):
	for i in catalogs:
		if not os.path.isdir(i):
			os.mkdir(i)

def xl_write(i, sheet, card):
		sheet.write(i, 0, card.type)
		sheet.write(i, 1, card.reason)
		sheet.write(i, 2, card.name)
		sheet.write(i, 3, card.surname)
		sheet.write(i, 4, card.inn)
		sheet.write(i, 5, card.phone)
		sheet.write(i, 6, card.pay_method)

def relocate_image(source, dest, card):
	if os.path.isfile(source):
		os.replace(source, dest)
		card.photo.name = dest
		card.save()


def MyCronJob():
	date = datetime.datetime.utcnow().strftime("%m%d")
	media = 'media/'


	#Create a catalog for new images
	makedirs(date, media, (media, media+date, media+f'{date}/st', media+f'{date}/sch'))

	loadout = Loadout.objects.filter(pub_date__gte=datetime.datetime.now().replace(hour=0,minute=0,second=0))

	#Load CardModels from a db
	cards = Card.objects.all()

	#Write data to an excel file and copy images
	st_book = xlwt.Workbook(encoding="utf-8")
	sch_book = xlwt.Workbook(encoding="utf-8")
	st_sheet = st_book.add_sheet("St")
	sch_sheet = sch_book.add_sheet("Sch")
	st_row = 0
	sch_row = 0
	st_zip = ZipFile(media + f"{date}/st.zip", 'w')
	sch_zip = ZipFile(media + f"{date}/sch.zip", 'w')
	st_path = f'{date}/st'
	sch_path = f'{date}/sch'
	for card in cards:
		if card.pub_date.strftime("%m%d") == date:
			if card.type == 'Школьная':
				xl_write(sch_row, sch_sheet, card)
				relocate_image(media+str(card.photo), media+sch_path+f'/{str(card.inn)+str(card.photo)[-4:]}', card)
				sch_zip.write(card.photo.name)
				sch_row+=1
				continue
			if card.type == 'Студенческая':
				xl_write(st_row, st_sheet, card)
				relocate_image(media+str(card.photo), media+st_path+f'/{str(card.inn)+str(card.photo)[-4:]}', card)
				st_zip.write(card.photo.name)
				st_row+=1
				continue
		break
	st_zip.close()
	sch_zip.close()
	st_book.save(media+st_path+'.xls')
	sch_book.save(media+sch_path+'.xls')

	if loadout.count() == 0:
		loadout = Loadout()
		loadout.st_xls=st_path+'.xls'
		loadout.sch_xls=sch_path+'.xls'
		loadout.st_images=st_path+'.zip'
		loadout.sch_images=sch_path+'.zip'
		loadout.save()

def check_inn(query, path):
	rb = openpyxl.load_workbook('media/' + str(path))
	names = rb.get_sheet_names()
	sheet = rb.get_sheet_by_name(names[0])
	for i in range(1, sheet.max_row):
		val = sheet.cell(row=i, column=1).value.lower()
		if 'iin' in val:
			if query == val[-12:]:
				return True

	return False
	
	

